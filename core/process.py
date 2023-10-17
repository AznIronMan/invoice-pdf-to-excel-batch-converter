import openpyxl
import pandas as pd
import pdfplumber
import json
import os
import re

from datetime import datetime
from dateutil import parser
from dotenv import load_dotenv
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Tuple, Union

from core.logger import zlog as log


load_dotenv()


def batch_convert(target_dir: Optional[str] = None) -> bool:
    try:
        if target_dir is None:
            target_dir = os.path.dirname(os.path.abspath(__file__))
    except Exception as e:
        error = f"Error getting target directory: {e}"
        log(error, "FATAL")
        raise Exception(error)

    for root, _, files in os.walk(target_dir):
        processed_dir = os.getenv('PROCESSED_DIR', 'processed')
        processed_folder = os.path.join(root, processed_dir)

        for pdf_file in files:
            if pdf_file.endswith('.pdf'):
                if not os.path.exists(processed_folder):
                    os.mkdir(processed_folder)

                pdf_path = os.path.join(root, pdf_file)

                excel_path = os.path.join(
                    processed_folder, pdf_file.replace('.pdf', '.xlsx'))

                try:
                    if pdf_to_excel(pdf_path, excel_path):
                        success = f"Processed {pdf_file} -> {excel_path}"
                        log(success, "INFO", True)
                    else:
                        if os.path.exists(excel_path):
                            kinda_success = f"Processed {pdf_file} -> {excel_path} but without formatting!"
                            log(kinda_success, "WARNING")
                        else:
                            raise Exception(
                                "Error processing pdf to excel!")
                except Exception as e:
                    error = f"Error processing {pdf_file} -> {e}"
                    log(error, "ERROR")
                    continue


def clean_currency(value: str) -> str:
    return re.sub(r'[^\d$.]', '', value)


def contains_email_address(line: str) -> bool:
    pattern = re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}')
    return bool(pattern.search(line))


def contains_phone_number(line: str) -> bool:
    pattern = re.compile(
        r'((\+?1\s*)?(\(\d{3}\)\s*|\d{3}[-.\s]?)\d{3}[-.\s]?\d{4})'
    )
    return bool(pattern.search(line))


def find_and_parse_date(lines: List[str]) -> Tuple[Dict, Union[int, None]]:
    try:
        date_pattern = (
            r"(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4}|\d{1,2}/\d{1,2}/\d{2,4}"
        )

        for index, line in enumerate(lines):
            for year in get_years_to_search():
                if str(year) in line or f"/{str(year)[2:]}" in line:
                    match = re.search(date_pattern, line)
                    if match:
                        date_str = match.group(0)
                        try:
                            date = parser.parse(date_str.strip(), fuzzy=True)
                            data = {'Date': date.strftime('%m/%d/%Y')}
                            return data, index
                        except ValueError:
                            print(f"Error parsing date: {date_str.strip()}")
        data = {'Date': datetime.now().strftime('%m/%d/%Y')}
        return data, None
    except Exception as e:
        error = f"Error finding and parsing date: {e}"
        log(error, "ERROR")
        return {}, None


def find_header_fill_index(lines: List[str], start_index: int) -> int:
    product_header = None

    for line in lines[start_index:]:
        if line.startswith("Product Description Cost per Item Qty Price"):
            product_header = lines.index(line)
            break
        elif line.startswith("Product Description"):
            product_header = lines.index(line)
            break
        elif line.startswith("Description Quantity Price Total Price"):
            product_header = lines.index(line)
            break

    return product_header


def format_excel(excel_path) -> bool:
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        header_fill = os.getenv('HEADER_FILL', json.dumps(
            ['4CAF50', '4CAF50', 'solid']))

        header_start_color, header_end_color, header_fill_type = json.loads(
            header_fill)

        green_fill = PatternFill(
            start_color=header_start_color,
            end_color=header_end_color,
            fill_type=header_fill_type
        )

        last_header_col = 0
        for cell in ws["1:1"]:
            if cell.value is not None:
                last_header_col = cell.column
                cell.fill = green_fill

        ws.auto_filter.ref = f"A1:{get_column_letter(last_header_col)}1"

        for col in range(1, last_header_col + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            for cell in ws[col_letter]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(excel_path)
        return True
    except Exception as e:
        error = f"Error formatting excel -> {excel_path}: {e}"
        log(error, "WARNING")
        return False


def get_years_to_search(year_range: int = 10) -> List[int]:
    current_year = datetime.now().year
    try:
        years_to_search = [current_year +
                           i for i in range(-year_range, year_range + 1)]
        years_to_search = sorted(
            years_to_search, key=lambda x: abs(current_year - x))
        return years_to_search
    except Exception as e:
        error = f"Error getting years to search: {e} -> returning current year!"
        log(error, "WARNING")
        return [current_year]


def is_city_state_zip_line(line: str) -> bool:
    pattern = re.compile(r'^[\w\s]+,\s*\w+\s+\d+')
    return bool(pattern.match(line))


def is_department_line(line: str) -> bool:
    if (
        not is_traditional_address(line) and
        not is_po_box_address(line) and
        not is_city_state_zip_line(line) and
        not contains_phone_number(line) and
        not contains_email_address(line) and
        not starts_with_invoice_or_purchase(line)
    ):
        return True
    else:
        return False


def is_po_box_address(line: str) -> bool:
    pattern = re.compile(r'P\.?O\.?\s*Box\s+\d+', re.IGNORECASE)
    return bool(pattern.match(line))


def is_traditional_address(line: str) -> bool:
    pattern = re.compile(r'^\d+\s[\w\s]+')
    return bool(pattern.match(line))


def isolate_email(line: str) -> str:
    pattern = re.compile(
        r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})'
    )
    match = pattern.search(line)
    return match.group(0) if match else None


def isolate_number(line: str) -> str:
    pattern = re.compile(
        r'((\+?1\s*)?(\(\d{3}\)\s*|\d{3}[-.\s]?)\d{3}[-.\s]?\d{4})'
    )
    match = pattern.search(line)
    return match.group(0) if match else None


def map_text_to_excel_columns(text: str) -> pd.DataFrame:
    try:
        lines = text.strip().split('\n')
        mapped_data = {}

        date_data, date_index = find_and_parse_date(lines)
        mapped_data.update(date_data)

        document_data, document_end_index = parse_main_section(
            lines, date_index + 1)
        mapped_data.update(document_data)

        invoice_data, invoice_po_end_index = parse_invoice_and_purchase_order(
            lines, document_end_index)
        mapped_data.update(invoice_data)

        product_data, freight_index = parse_products(
            lines, invoice_po_end_index)
        for i, product in enumerate(product_data):
            for key, value in product.items():
                mapped_data[f"{key}_{i}"] = value

        freight_data = parse_freight(lines, freight_index)
        mapped_data.update(freight_data)

        return pd.DataFrame([mapped_data])
    except Exception as e:
        error = f"Error mapping text to excel columns: {e}"
        log(error, "CRITICAL")
        return pd.DataFrame()


def parse_main_section(lines: List[str], start_index: int) -> Tuple[Dict, int]:
    data = {}
    current_index = start_index
    contact_count = 0
    product_header = find_header_fill_index(lines, start_index)

    while current_index < product_header:
        line = lines[current_index].strip()

        address_1, success = parse_address_1(line)
        if success:
            data['Address 1'] = address_1
            current_index += 1
            line = lines[current_index].strip()

        if address_1:
            address_2, success = parse_address_2(line)
            if success:
                data['Address 2'] = address_2
                current_index += 1
                line = lines[current_index].strip()

        city_state_zip, success = parse_citystatezip(line)
        if success:
            data['City, State, Zip'] = city_state_zip
            current_index += 1
            line = lines[current_index].strip()

        contact, success = parse_contact(line)
        while success:
            contact_count += 1
            data[f'Contact {contact_count}'] = contact
            current_index += 1
            if current_index < product_header:
                line = lines[current_index].strip()
                contact, success = parse_contact(line)
            else:
                break

        phone, success = parse_phone(line)
        if success:
            data['Phone'] = phone
            current_index += 1
            if current_index < len(lines):
                line = lines[current_index].strip()

        email, success = parse_email(line)
        if success:
            data['Email'] = email
            current_index += 1
            if current_index < len(lines):
                line = lines[current_index].strip()
        current_index += 1

    return data, current_index


def parse_address_1(line: str) -> Tuple[str, bool]:
    try:
        if isinstance(line, str) and not None:
            if ((is_traditional_address(line) or is_po_box_address(line))) and (
                not is_city_state_zip_line(line) or not contains_phone_number(line) or
                    not contains_email_address(line) or not starts_with_invoice_or_purchase(line)):
                return line.strip(), True
        return None, False
    except Exception as e:
        error = f"Error parsing address_1: {e}"
        log(error, "ERROR")
        return None, False


def parse_address_2(line: str) -> Tuple[str, bool]:
    try:
        if isinstance(line, str) and not None:
            if (not is_traditional_address(line) and not is_po_box_address(line) and
                not is_city_state_zip_line(line) and not contains_phone_number(line) and
                    not contains_email_address(line) and not starts_with_invoice_or_purchase(line)):
                return line.strip(), True
        return None, False
    except Exception as e:
        error = f"Error parsing address_2: {e}"
        log(error, "ERROR")
        return None, False


def parse_citystatezip(line: str) -> Tuple[str, bool]:
    try:
        if isinstance(line, str) and not None:
            if is_city_state_zip_line(line):
                return line.strip(), True
        return None, False
    except Exception as e:
        error = f"Error parsing citystatezip: {e}"
        log(error, "ERROR")
        return None, False


def parse_contact(line: str) -> Tuple[str, bool]:
    try:
        if isinstance(line, str) and not None:
            if (not contains_email_address(line) and not contains_phone_number(line) and
                    not starts_with_invoice_or_purchase(line)):
                return line.strip(), True
        return None, False
    except Exception as e:
        error = f"Error parsing contact: {e}"
        log(error, "ERROR")
        return None, False


def parse_department(line: str) -> Tuple[Optional[str], bool]:
    try:
        if is_department_line(line):
            return line.strip(), True
        else:
            return None, False
    except Exception as e:
        error = f"Error parsing department: {e}"
        log(error, "ERROR")
        return None, False


def parse_email(line: str) -> Tuple[str, bool]:
    try:
        if contains_email_address(line):
            return isolate_email(line), True
        return None, False
    except Exception as e:
        error = f"Error parsing email: {e}"
        log(error, "ERROR")
        return None, False


def parse_email(line: str) -> Tuple[str, bool]:
    try:
        if "@" in line:
            email = line.strip().split()[-1]
            return email, True
        return None, False
    except Exception as e:
        error = f"Error parsing email: {e}"
        log(error, "ERROR")
        return None, False


def parse_freight(lines: List[str], start_index: int) -> Dict:
    try:
        data = {}
        for line in lines[start_index:]:
            if line.startswith("Freight"):
                if ":" in line:
                    data['Freight'] = clean_currency(
                        line.split(":")[1].strip())
                    break
                else:
                    data['Freight'] = clean_currency(
                        line.split("Freight")[1].strip())
                    break
        return data
    except Exception as e:
        error = f"Error parsing freight: {e}"
        log(error, "ERROR")
        return {}


def parse_invoice_and_purchase_order(lines: List[str], start_index: int) -> Dict:
    try:
        data = {}
        invoice, po = None, None
        invoice_line, po_line = None, None
        for line in lines[start_index:]:
            if line.startswith("Invoice") and not invoice:
                invoice = line.split(":")[1].strip()
                data['Invoice'] = invoice
                invoice_line = lines.index(line)
            if line.startswith("Purchase") and not po:
                po = line.split(":")[1].strip()
                data['Purchase Order'] = po
                po_line = lines.index(line)
            if invoice and po:
                break
        return_line = ((invoice_line if invoice_line > po_line else po_line)
                       if invoice_line and po_line else start_index)
        return data, return_line
    except Exception as e:
        error = f"Error parsing invoice and purchase order: {e}"
        log(error, "ERROR")
        return {}


def parse_phone(line: str) -> Tuple[Dict[str, str], bool]:
    try:
        tel_array = os.getenv(
            'MAIN_PHONE', json.dumps(['Tel', 'Main', 'Home', 'Office', 'Phone', 'Telephone']))
        cell_array = os.getenv('CELL_PHONE', json.dumps(
            ['Cell', 'Mobile', 'iPhone']))
        keys = {"Tel": json.loads(tel_array), "Cell": json.loads(cell_array)}
        phone_data = {}
        for key, prefixes in keys.items():
            if any(line.startswith(prefix) for prefix in prefixes):
                match = re.search(r'\d{3}[-.\s]?\d{3}[-.\s]?\d{4}', line)
                if match:
                    phone_data[key] = match.group(0)
                    return phone_data, True
        return None, False
    except Exception as e:
        error = f"Error parsing phone: {e}"
        log(error, "ERROR")
        return None, False


def parse_products(lines: List[str], start_index: int) -> List[Dict]:
    data = []
    all_products = False
    product_header = None
    try:
        product_header = find_header_fill_index(lines, start_index)

        if product_header is None:
            log("Product header not found", "ERROR")
            return [], None

        while not all_products:
            for line in lines[product_header+1:]:
                product_data = line.split(" ")
                product_data = list(filter(None, product_data))
                if len(product_data) >= 4:
                    product_name, per_price, quantity, total_price = " ".join(
                        product_data[:-3]), product_data[-3], product_data[-2], product_data[-1]
                    data.append({
                        'Product_Description': product_name,
                        'Price_Per_Product': clean_currency(per_price),
                        'Quantity': clean_currency(quantity),
                        'Total_Price': clean_currency(total_price)
                    })
                if line.startswith("Freight"):
                    freight_line = lines.index(line)
                    all_products = True
                    break
        return data, freight_line
    except Exception as e:
        error = f"Error parsing products: {e}"
        log(error, "ERROR")
        return [], None


def pdf_to_excel(pdf_path, excel_path) -> bool:
    try:
        with pdfplumber.open(pdf_path) as pdf:
            page = pdf.pages[0]
            text = page.extract_text()

        df = map_text_to_excel_columns(text)

        if df.empty:
            raise Exception("Error mapping text to excel columns!")
        else:
            df.to_excel(excel_path, index=False)
            if format_excel(excel_path):
                return True
            else:
                raise Exception("Error formatting excel!")
    except Exception as e:
        error = f"Error converting pdf [{pdf_path}] -> {excel_path}: {e}"
        log(error, "ERROR")
        return False


def starts_with_invoice_or_purchase(line: str) -> bool:
    lower_line = line.lower()
    return lower_line.startswith(('invoice', 'purchase'))
